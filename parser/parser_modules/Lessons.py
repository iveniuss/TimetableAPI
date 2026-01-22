from fnmatch import fnmatch
from parser.parser_modules.Lesson import Lesson
from openpyxl.worksheet.worksheet import Worksheet


TL_pat = "* ?.?. (*)"

WEEKDAYS = ("Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота")
IGNORE = ("None", "Английский язык", "Военная кафедра") # Пары с такими словами будут игнорироваться


def _format_date_time(start_time, end_time, date):
    if len(start_time) == 4:
        start_time = "0" + start_time
    if len(end_time) == 4:
        end_time = "0" + end_time
    start_time += ":00"
    end_time += ":00"
    day, month, year = date.split(".")
    date = f"{year}-{month}-{day}"
    return f"{date}T{start_time}", f"{date}T{end_time}"


def _split_teacher_location(description):
    open_scope = description.rfind("(")
    close_scope = description.rfind(")")
    if open_scope != -1:
        location = description[open_scope + 1 : close_scope]
        teacher = description[:open_scope]
    else:
        location = ""
        teacher = description

    return teacher, location





class Lessons(object):

    def __init__(self, group: dict):
        self.lessons = []
        self.group_info = group

    # Получение номера столбца нужной группы
    def _get_group_column(self, sheet:Worksheet):
        row = 3
        
        for column in range(3,sheet.max_column):
            if sheet.cell(row=row, column=column).value == self.group_info.get('name'):
                return column
        return -1

    # Добавление всех пар заданной группы с листа
    def add_lessons_from_sheet(self, sheet):
        column = self._get_group_column(sheet)
        if column < 1:
            return
        for row in range(4, sheet.max_row):
            lesson_cell = str(
                sheet.cell(row=row, column=column).value
            )
            date_cell = str(sheet.cell(row=row, column=1).value)
            time_cell = str(sheet.cell(row=row, column=2).value)[3:]
            try:
                weekday, date = date_cell.split("\n")
                start_time, end_time = time_cell.split("-")
            except:
                continue

            for l in self._parse_lessons(lesson_cell):
                t = True
                for s in IGNORE:
                    if s in l.name:
                        t = False
                
                if not t: continue
                start, end = _format_date_time(start_time, end_time, date)
                l.add_date(start, end)
                self.lessons.append(l)

    # Разделение текста в ячейке на список пар
    def _parse_lessons(self, cell_value: str) -> list[Lesson]:
        cell_list = list(filter(None,cell_value.split("\n")))


        # Определение паттерна в ячейке
        order = ""
        for s in cell_list:
            if fnmatch(s, TL_pat):
                order+="D" # Описание (Имя преподавателя + аудитория)
            elif fnmatch(s, "http*"):
                order+="L" # Ссылка на онлайн пару
            else:
                order+="N" # Название пары

        lessons = []

        # Разбиение текста в ячейке на части
        if order == "ND":
            tchr, loc = _split_teacher_location(cell_list[1])
            lessons.append(Lesson(cell_list[0], tchr, loc))
        elif order == "NDND":
            tchr, loc = _split_teacher_location(cell_list[1])
            lessons.append(Lesson(cell_list[0], tchr, loc))
            tchr, loc = _split_teacher_location(cell_list[3])
            lessons.append(Lesson(cell_list[2], tchr, loc))
        elif order == "NDL":
            tchr, loc = _split_teacher_location(cell_list[1])
            lessons.append(Lesson(cell_list[0], tchr, loc, link=cell_list[2]))
        elif order == "NDD":
            tchr, loc = _split_teacher_location(cell_list[1])
            lessons.append(Lesson(cell_list[0], tchr, loc))

            tchr, loc = _split_teacher_location(cell_list[2])
            lessons.append(Lesson(cell_list[0], tchr, loc))
        elif order == "N":
            lessons.append(Lesson(cell_list[0]))
        else:
            mult_lessons = []
            index = 0
            flag = False
            while index < len(cell_list)-1:
                if order[index:index+3] == "NDL":
                    tchr, loc = _split_teacher_location(cell_list[index + 1])
                    mult_lessons.append(Lesson(cell_list[index], tchr, loc, link=cell_list[index + 2]))
                    index+=3
                elif order[index:index+2] == "ND":
                    tchr, loc = _split_teacher_location(cell_list[index+1])
                    mult_lessons.append(Lesson(cell_list[index], tchr, loc))
                    index+=2
                else:
                    flag = True
                    break

            if not flag:
                lessons += mult_lessons

        # Разделение на подгруппы
        for lesson in lessons.copy():
            if (',' in lesson.location
                    and "НИС" not in lesson.name
                    and "Коммуникационный менеджмент" not in lesson.name
                    and "Организационное поведение" not in lesson.name):
                lesson.group = lesson.location[-1]
                lesson.location = lesson.location[:lesson.location.find(',')]
            else:
                lesson.group = self.group_info['subgroups'][0]
                copy = lesson.copy()
                copy.group = self.group_info['subgroups'][1]
                lessons.append(copy)
        
        return lessons

    def get_lessons_dict(self, prev_timetable) -> dict[str, list[Lesson]]:
        only_in_prev = [
            prev_lesson
            for prev_lesson in prev_timetable
            if not any(prev_lesson == actual_lesson
                for actual_lesson in self.lessons
            )
        ]

        only_in_actual = [
            actual_lesson
            for actual_lesson in self.lessons
            if not any(actual_lesson == prev_event
                for prev_event in prev_timetable
            )
        ]

        return {"to_add": only_in_actual, "to_del": only_in_prev}
